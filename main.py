from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QFileDialog

from pytube.innertube import _default_clients
from pytube import cipher
import re

from openpyxl import load_workbook

from pytubefix import YouTube

import requests

from rutube import Rutube

from openpyxl.drawing.image import Image

from urllib.parse import urlparse

from fake_useragent import UserAgent

# _default_clients["ANDROID"]["context"]["client"]["clientVersion"] = "19.08.35"
# _default_clients["IOS"]["context"]["client"]["clientVersion"] = "19.08.35"
# _default_clients["ANDROID_EMBED"]["context"]["client"]["clientVersion"] = "19.08.35"
# _default_clients["IOS_EMBED"]["context"]["client"]["clientVersion"] = "19.08.35"
# _default_clients["IOS_MUSIC"]["context"]["client"]["clientVersion"] = "6.41"
# _default_clients["ANDROID_MUSIC"] = _default_clients["ANDROID_CREATOR"]
#
#
# def get_throttling_function_name(js: str) -> str:
#     """Extract the name of the function that computes the throttling parameter.
#
#     param str js:
#         The contents of the base.js asset file.
#     :rtype: str
#     :returns:
#         The name of the function used to compute the throttling parameter.
#     """
#     function_patterns = [
#         r'a\.[a-zA-Z]\s*&&\s*\([a-z]\s*=\s*a\.get\("n"\)\)\s*&&\s*'
#         r'\([a-z]\s*=\s*([a-zA-Z0-9$]+)(\[\d+\])?\([a-z]\)',
#         r'\([a-z]\s*=\s*([a-zA-Z0-9$]+)(\[\d+\])\([a-z]\)',
#     ]
#     # logger.debug('Finding throttling function name')
#     for pattern in function_patterns:
#         regex = re.compile(pattern)
#         function_match = regex.search(js)
#         if function_match:
#             # logger.debug("finished regex search, matched: %s", pattern)
#             if len(function_match.groups()) == 1:
#                 return function_match.group(1)
#             idx = function_match.group(2)
#             if idx:
#                 idx = idx.strip("[]")
#                 array = re.search(
#                     r'var {nfunc}\s*=\s*(\[.+?\]);'.format(
#                         nfunc=re.escape(function_match.group(1))),
#                     js
#                 )
#                 if array:
#                     array = array.group(1).strip("[]").split(",")
#                     array = [x.strip() for x in array]
#                     return array[int(idx)]
#
#
# cipher.get_throttling_function_name = get_throttling_function_name




class video(QtCore.QThread):
    load_finshed = QtCore.pyqtSignal(object)

    def __init__(self, path):
        super().__init__()
        self.pathXlsx = path


    def run(self):
        wb = load_workbook(self.pathXlsx)
        for sheet in wb.sheetnames:
            worksheet = wb[sheet]
            QtWidgets.QApplication.processEvents()
            for row in worksheet.iter_rows(values_only=True):
                QtWidgets.QApplication.processEvents()
                for cell in row:
                    QtWidgets.QApplication.processEvents()
                    if isinstance(cell, str):
                        cell_normal = cell
                        try:
                            parseUrl = urlparse(cell_normal)
                            id = parseUrl.path.split("/")[-1]
                            url = "https://youtube.com/" + id
                            print(url)
                            yt = YouTube(url)
                            video = yt.streams.get_lowest_resolution()
                            print("download start")
                            try:
                                video.download()
                            except Exception as e:
                                print(e)


                            def clean_filename(title):
                                print("name cleaning")
                                return re.sub(r"[^\w\s-]", "", title).strip().replace(" ", "_")


                            prew = yt.thumbnail_url

                            response = requests.get(prew)
                            img_data = response.content

                            # Использование обработанного названия видео в имени файла
                            cleaned_title = clean_filename(video.title)
                            filename = f"{cleaned_title}.png"
                            print(filename)

                            with open(filename, "xb") as file:
                                file.write(img_data)

                            self.finished.emit()
                            QtWidgets.QApplication.processEvents()
                            print("video is downloaded")
                        except Exception as e:
                            print(e)
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.pathToFile = ""
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.Path = QtWidgets.QTextEdit(parent=self.centralwidget)
        self.Path.setGeometry(QtCore.QRect(90, 140, 591, 31))
        self.Path.setStyleSheet("background-color: rgb(244, 216, 255);")
        self.Path.setObjectName("Path")
        self.startDownloadButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.startDownloadButton.setGeometry(QtCore.QRect(530, 240, 141, 28))
        self.startDownloadButton.setStyleSheet("background-color: rgb(68, 255, 0);")
        self.startDownloadButton.setObjectName("startDownloadButton")
        self.startDownloadButton.clicked.connect(self.startDownload)
        self.stopDonwloadButtob = QtWidgets.QPushButton(parent=self.centralwidget)
        self.stopDonwloadButtob.setGeometry(QtCore.QRect(100, 240, 161, 28))
        self.stopDonwloadButtob.setStyleSheet("background-color: rgb(255, 0, 0);")
        self.stopDonwloadButtob.setObjectName("stopDonwloadButtob")
        self.stopDonwloadButtob.clicked.connect(self.open_file_dialog)
        self.checkBox = QtWidgets.QCheckBox(parent=self.centralwidget)
        self.checkBox.setGeometry(QtCore.QRect(600, 200, 311, 20))
        self.checkBox.setObjectName("checkBox")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.progressBar = QtWidgets.QProgressBar(parent=MainWindow)
        self.progressBar.setGeometry(QtCore.QRect(300, 255, 221, 21))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")



        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.Path.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:7.8pt; font-weight:400; font-style:normal;\">\n"
f"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt;\">{self.pathToFile}</span></p></body></html>"))
        self.startDownloadButton.setText(_translate("MainWindow", "Начать скачивание"))
        self.stopDonwloadButtob.setText(_translate("MainWindow", "Выбрать файл"))


    def startDownload(self):
        self.thread = video(self.pathToFile)
        self.thread.start()


    def open_file_dialog(self):
        file_name, _ = QFileDialog.getOpenFileName(parent=None, caption="Выбрать файл")

        self.pathToFile = file_name


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())


